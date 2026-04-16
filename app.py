"""
Кладка Хаб — Flask приложение

Архитектура хранения данных:
  Excel  — для людей (загрузка от заказчика / выгрузка заказчику)
  JSON   — для данных (внутреннее хранение, git-версионирование)
  MD     — для ИИ (база знаний, контекст для Claude API)
"""
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import os
from functools import wraps
from datetime import datetime
import io

app = Flask(__name__)
app.secret_key = 'kladka-hub-secret-2026'

# ── Пользователи ───────────────────────────────────────────────────────────────
USERS = {
    'veduschiy': {'password': 'kh2026',  'role': 'lead',    'name': 'Ведущий инженер'},
    'inzhener':  {'password': 'kh2026r', 'role': 'pricing', 'name': 'Инженер расценки'}
}

# ── Пути к файлам ──────────────────────────────────────────────────────────────
DATA_DIR         = 'data'
VOR_JSON         = os.path.join(DATA_DIR, 'vor_kholodov.json')
SPRAVOCHNIK_JSON = os.path.join(DATA_DIR, 'spravochnik.json')
MAPPING_JSON     = os.path.join(DATA_DIR, 'mapping.json')


@app.context_processor
def inject_now():
    return {'now': datetime.now().strftime('%d.%m.%Y')}


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── JSON: чтение ───────────────────────────────────────────────────────────────

def load_vor():
    """Читает ВОР из JSON. Возвращает список позиций."""
    if not os.path.exists(VOR_JSON):
        return []
    with open(VOR_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get('items', [])


def load_spravochnik():
    """Читает справочник цен из JSON. Возвращает список материалов."""
    if not os.path.exists(SPRAVOCHNIK_JSON):
        return []
    with open(SPRAVOCHNIK_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get('items', [])


def load_mapping():
    """Читает маппинг позиций ВОР → материал из JSON."""
    if not os.path.exists(MAPPING_JSON):
        return {}
    with open(MAPPING_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


# ── JSON: запись ───────────────────────────────────────────────────────────────

def save_vor(items, meta=None):
    """Сохраняет ВОР в JSON."""
    data = {
        "meta": meta or {
            "tender": "ЖК Холодов",
            "section": "Кладка",
            "loaded_at": datetime.now().strftime('%d.%m.%Y'),
            "loaded_by": session.get('user', 'system')
        },
        "items": items
    }
    with open(VOR_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_spravochnik(items):
    """Сохраняет справочник цен в JSON."""
    data = {
        "meta": {
            "section": "Кладка",
            "updated_at": datetime.now().strftime('%d.%m.%Y'),
            "updated_by": session.get('user', 'system')
        },
        "items": items
    }
    with open(SPRAVOCHNIK_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_mapping(mapping):
    """Сохраняет маппинг в JSON."""
    with open(MAPPING_JSON, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)


# ── Excel → JSON конвертеры ────────────────────────────────────────────────────

def excel_to_vor_json(filepath):
    """Конвертирует Excel ВОР в JSON. Возвращает список позиций."""
    wb = load_workbook(filepath)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            items.append({
                'num':          row[0],
                'naimenovanie': str(row[1]) if row[1] else '',
                'ed_izm':       str(row[2]) if row[2] else '',
                'kolvo':        float(row[3]) if row[3] else 0.0
            })
    return items


def excel_to_spravochnik_json(filepath):
    """Конвертирует Excel справочника в JSON. Возвращает список материалов."""
    wb = load_workbook(filepath)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            date_val = row[4]
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%d.%m.%Y')
            else:
                date_str = str(date_val) if date_val else datetime.now().strftime('%d.%m.%Y')
            items.append({
                'nomenclatura': str(row[0]),
                'ed_izm':       str(row[1]) if row[1] else '',
                'cena':         float(row[2]) if row[2] else 0.0,
                'valuta':       str(row[3]) if row[3] else 'RUB',
                'data':         date_str
            })
    return items


# ── JSON → Excel конвертер (для экспорта) ──────────────────────────────────────

def build_export_excel():
    """Строит Excel-файл для заказчика из JSON данных."""
    vor_items    = load_vor()
    spravochnik  = load_spravochnik()
    mapping      = load_mapping()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Кладка — Расчёт'

    # Заголовок
    ws.merge_cells('A1:G1')
    ws['A1'] = f'ЖК Холодов | Раздел: Кладка | Расчёт от {datetime.now().strftime("%d.%m.%Y")}'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = PatternFill(fill_type='solid', fgColor='1a3a2a')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['№', 'Наименование по ВОР', 'Ед.изм', 'Кол-во', 'Материал', 'Цена за ед., руб', 'Итого, руб']
    ws.append(headers)
    for cell in ws[2]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='2d6a4f')
        cell.alignment = Alignment(horizontal='center')

    total = 0.0
    price_map = {s['nomenclatura']: s['cena'] for s in spravochnik}

    for item in vor_items:
        key      = str(item['num'])
        material = mapping.get(key, {}).get('material', '')
        cena     = price_map.get(material, 0.0) if material else 0.0
        itogo    = cena * item['kolvo']
        total   += itogo
        ws.append([item['num'], item['naimenovanie'], item['ed_izm'],
                   item['kolvo'], material, cena, itogo])

    # Итого
    tr = ws.max_row + 1
    ws.cell(tr, 6, 'ИТОГО:').font = Font(bold=True)
    ws.cell(tr, 7, total).font    = Font(bold=True)

    for i, w in enumerate([5, 45, 8, 10, 25, 18, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


# ── Расчёт итогов (общая функция) ──────────────────────────────────────────────

def calc_result():
    """Возвращает список позиций с ценами и общий итог."""
    vor_items   = load_vor()
    spravochnik = load_spravochnik()
    mapping     = load_mapping()
    price_map   = {s['nomenclatura']: s['cena'] for s in spravochnik}

    result = []
    for item in vor_items:
        key      = str(item['num'])
        material = mapping.get(key, {}).get('material', '')
        cena     = price_map.get(material, 0.0) if material else 0.0
        itogo    = cena * item['kolvo']
        result.append({**item, 'material': material, 'cena': cena, 'itogo': itogo})

    total = sum(r['itogo'] for r in result)
    return result, total


# ── Маршруты: авторизация ──────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username in USERS and USERS[username]['password'] == password:
            session['user'] = username
            session['role'] = USERS[username]['role']
            session['name'] = USERS[username]['name']
            return redirect(url_for('dashboard'))
        error = 'Неверный логин или пароль'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Маршруты: дашборд ──────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')


# ── Маршруты: ВОР ─────────────────────────────────────────────────────────────

@app.route('/tender/kholodov/kladka')
@login_required
def vor_kladka():
    result, total = calc_result()
    spravochnik   = load_spravochnik()
    return render_template('vor.html', items=result, spravochnik=spravochnik, total=total)


@app.route('/vor/upload', methods=['POST'])
@login_required
def vor_upload():
    """Загрузка Excel ВОР → конвертация в JSON."""
    if 'file' in request.files:
        f = request.files['file']
        if f.filename:
            tmp = os.path.join(DATA_DIR, '_tmp_vor.xlsx')
            f.save(tmp)
            items = excel_to_vor_json(tmp)
            save_vor(items)
            os.remove(tmp)
    return redirect(url_for('vor_kladka'))


@app.route('/vor/mapping', methods=['POST'])
@login_required
def vor_mapping():
    """Сохраняет привязку позиции ВОР к материалу."""
    data     = request.json
    pos_num  = str(data.get('num'))
    material = data.get('material', '')

    mapping = load_mapping()
    mapping[pos_num] = {'material': material}
    save_mapping(mapping)

    spravochnik = load_spravochnik()
    vor_items   = load_vor()
    price_map   = {s['nomenclatura']: s['cena'] for s in spravochnik}

    kolvo = next((i['kolvo'] for i in vor_items if str(i['num']) == pos_num), 0.0)
    cena  = price_map.get(material, 0.0) if material else 0.0
    itogo = cena * kolvo

    return jsonify({
        'success':    True,
        'cena':       cena,
        'itogo':      itogo,
        'cena_fmt':   f'{cena:,.0f}'.replace(',', ' '),
        'itogo_fmt':  f'{itogo:,.0f}'.replace(',', ' ')
    })


# ── Маршруты: справочник ───────────────────────────────────────────────────────

@app.route('/spravochnik')
@login_required
def spravochnik():
    items = load_spravochnik()
    return render_template('spravochnik.html', items=items)


@app.route('/spravochnik/upload', methods=['POST'])
@login_required
def spravochnik_upload():
    """Загрузка Excel справочника → конвертация в JSON."""
    if 'file' in request.files:
        f = request.files['file']
        if f.filename:
            tmp = os.path.join(DATA_DIR, '_tmp_spr.xlsx')
            f.save(tmp)
            items = excel_to_spravochnik_json(tmp)
            save_spravochnik(items)
            os.remove(tmp)
    return redirect(url_for('spravochnik'))


@app.route('/spravochnik/update', methods=['POST'])
@login_required
def spravochnik_update():
    """Обновление цены прямо на сайте."""
    if session.get('role') != 'pricing':
        return jsonify({'error': 'Нет прав'}), 403

    data  = request.json
    items = load_spravochnik()
    today = datetime.now().strftime('%d.%m.%Y')

    for item in items:
        if item['nomenclatura'] == data['nomenclatura']:
            item['cena'] = float(data['cena'])
            item['data'] = today
            break

    save_spravochnik(items)
    return jsonify({'success': True, 'data': today})


# ── Маршруты: экспорт ─────────────────────────────────────────────────────────

@app.route('/export')
@login_required
def export():
    """JSON данные → Excel для заказчика."""
    wb     = build_export_excel()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'kladka_kholodov_{datetime.now().strftime("%d%m%Y")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    app.run(debug=True, port=5000)
