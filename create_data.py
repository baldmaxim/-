"""
Запустите этот файл один раз — он создаст тестовые Excel-файлы:
  - data/spravochnik.xlsx  (справочник цен — 3 материала)
  - data/vor_kholodov.xlsx (ВОР ЖК Холодов — 3 позиции)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

os.makedirs('data', exist_ok=True)

# ── Справочник цен ─────────────────────────────────────────────────────────────
wb1 = Workbook()
ws1 = wb1.active
ws1.title = 'Справочник'

headers = ['Номенклатура', 'Ед.изм', 'Цена', 'Валюта', 'Дата']
ws1.append(headers)
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(fill_type='solid', fgColor='2d6a4f')
    cell.alignment = Alignment(horizontal='center')

materials = [
    ['Кирпич 120 мм',  'м2', 1200, 'RUB', '15.04.2026'],
    ['Кирпич 250 мм',  'м2', 1800, 'RUB', '15.04.2026'],
    ['СКЦ 80 мм',      'м2',  900, 'RUB', '15.04.2026'],
]
for row in materials:
    ws1.append(row)

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 14

wb1.save('data/spravochnik.xlsx')
print('OK: data/spravochnik.xlsx sozdan')

# ── ВОР ЖК Холодов — Кладка ───────────────────────────────────────────────────
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'ВОР Кладка'

headers2 = ['№', 'Наименование', 'Ед.изм', 'Кол-во']
ws2.append(headers2)
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(fill_type='solid', fgColor='1a3a2a')
    cell.alignment = Alignment(horizontal='center')

vor_rows = [
    [1, 'Устройство кирпичных стен 120 мм',               'м2', 60],
    [2, 'Устройство кирпичных стен 250 мм',               'м2', 74],
    [3, 'Устройство перегородок из камня СКЦ толщ. 80 мм','м2', 1000],
]
for row in vor_rows:
    ws2.append(row)

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 52
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10

wb2.save('data/vor_kholodov.xlsx')
print('OK: data/vor_kholodov.xlsx sozdan')

print('\nGotovo! Mozhno zapuskat: python app.py')
